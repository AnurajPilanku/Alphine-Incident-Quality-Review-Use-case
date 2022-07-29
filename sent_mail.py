#Anuraj Pilanku
#Alphine usecase

import re
import os
import sys
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email.mime.image import MIMEImage
from email.mime.application import MIMEApplication
from os.path import basename
import openpyxl
import pandas as pd
#open txt and asign the data to a variable here instead of x
#x='''<br /><b>Incident</b> - <a href=https://gsm.mmm.com/ITSM/Inc_Detail.aspx?ID=IM9456199 >IM9456199</a><br /> <b>Priority</b> - Average<br /><b>Alert</b> - no update in incident log since 32 day(s)Alert_delimiter<br /><b>Required action</b> - please update incident log according to <a href=https://skydrive3m.sharepoint.com/sites/GOS/GOS%20Wiki/GOS%20Incident%20Support%20Process%20Guidelines/GOS%20Support%20Checklist.aspx?WikiPageMode=Edit&InitialTabId=Ribbon.EditingTools.CPEditTab&VisibilityContext=WSSWikiPage>3M Standards</a><br /><b>Required action</b> - review IM status 'Pending client' and proceed according to <a href=https://skydrive3m.sharepoint.com/sites/PolCent/Policy%20Central/forms/DispForm.aspx?ID=10703&Source=https://skydrive3m.sharepoint.com/sites/PolCent&InitialTabID=Ribbon.Read%20>IT GSOP-040</a><br /> * * * * * * * * * * * * * * * * * * * *<br /><b>Incident</b> - <a href=https://gsm.mmm.com/ITSM/Inc_Detail.aspx?ID=IM9356629 >IM9356629</a><br /> <b>Priority</b> - Average<br /><b>Alert</b> - no update in incident log since 47 day(s)Alert_delimiter<br /><b>Required action</b> - please update incident log according to <a href=https://skydrive3m.sharepoint.com/sites/GOS/GOS%20Wiki/GOS%20Incident%20Support%20Process%20Guidelines/GOS%20Support%20Checklist.aspx?WikiPageMode=Edit&InitialTabId=Ribbon.EditingTools.CPEditTab&VisibilityContext=WSSWikiPage>3M Standards</a><br /><b>Required action</b> - review IM status 'Pending client' and proceed according to <a href=https://skydrive3m.sharepoint.com/sites/PolCent/Policy%20Central/forms/DispForm.aspx?ID=10703&Source=https://skydrive3m.sharepoint.com/sites/PolCent&InitialTabID=Ribbon.Read%20>IT GSOP-040</a><br /> * * * * * * * * * * * * * * * * * * * *<br /><b>Incident</b> - <a href=https://gsm.mmm.com/ITSM/Inc_Detail.aspx?ID=IM9505566 >IM9505566</a><br /> <b>Priority</b> - High<br /><b>Alert</b> - no update in incident log since 27 day(s)Alert_delimiter<br /><b>Required action</b> - please update incident log according to <a href=https://skydrive3m.sharepoint.com/sites/GOS/GOS%20Wiki/GOS%20Incident%20Support%20Process%20Guidelines/GOS%20Support%20Checklist.aspx?WikiPageMode=Edit&InitialTabId=Ribbon.EditingTools.CPEditTab&VisibilityContext=WSSWikiPage>3M Standards</a><br /><b>Required action</b> - review IM status 'Pending client' and proceed according to <a href=https://skydrive3m.sharepoint.com/sites/PolCent/Policy%20Central/forms/DispForm.aspx?ID=10703&Source=https://skydrive3m.sharepoint.com/sites/PolCent&InitialTabID=Ribbon.Read%20>IT GSOP-040</a><br /> * * * * * * * * * * * * * * * * * * * *<br /><b>Incident</b> - <a href=https://gsm.mmm.com/ITSM/Inc_Detail.aspx?ID=IM9186827 >IM9186827</a><br /> <b>Priority</b> - Average<br /><b>Alert</b> - no update in incident log since 47 day(s)<br /><b>Required action</b> - please update incident log according to <a href=https://skydrive3m.sharepoint.com/sites/GOS/GOS%20Wiki/GOS%20Incident%20Support%20Process%20Guidelines/GOS%20Support%20Checklist.aspx?WikiPageMode=Edit&InitialTabId=Ribbon.EditingTools.CPEditTab&VisibilityContext=WSSWikiPage>3M Standards</a><br /><b>Required action</b> - review IM status 'Pending client' and proceed according to <a href=https://skydrive3m.sharepoint.com/sites/PolCent/Policy%20Central/forms/DispForm.aspx?ID=10703&Source=https://skydrive3m.sharepoint.com/sites/PolCent&InitialTabID=Ribbon.Read%20>IT GSOP-040</a><br /> * * * * * * * * * * * * * * * * * * * *<br /><b>Incident</b> - <a href=https://gsm.mmm.com/ITSM/Inc_Detail.aspx?ID=IM9454905 >IM9454905</a><br /> <b>Priority</b> - Average<br /><b>Alert</b> - no update in incident log since 34 day(s)<br /><b>Required action</b> - please update incident log according to <a href=https://skydrive3m.sharepoint.com/sites/GOS/GOS%20Wiki/GOS%20Incident%20Support%20Process%20Guidelines/GOS%20Support%20Checklist.aspx?WikiPageMode=Edit&InitialTabId=Ribbon.EditingTools.CPEditTab&VisibilityContext=WSSWikiPage>3M Standards</a><br /><b>Required action</b> - review IM status 'Pending client' and proceed according to <a href=https://skydrive3m.sharepoint.com/sites/PolCent/Policy%20Central/forms/DispForm.aspx?ID=10703&Source=https://skydrive3m.sharepoint.com/sites/PolCent&InitialTabID=Ribbon.Read%20>IT GSOP-040</a><br /> * * * * * * * * * * * * * * * * * * * *<br /><b>Incident</b> - <a href=https://gsm.mmm.com/ITSM/Inc_Detail.aspx?ID=IM9266860 >IM9266860</a><br /> <b>Priority</b> - High<br /><b>Alert</b> - no update in incident log since 74 day(s)<br /><b>Required action</b> - please update incident log according to <a href=https://skydrive3m.sharepoint.com/sites/GOS/GOS%20Wiki/GOS%20Incident%20Support%20Process%20Guidelines/GOS%20Support%20Checklist.aspx?WikiPageMode=Edit&InitialTabId=Ribbon.EditingTools.CPEditTab&VisibilityContext=WSSWikiPage>3M Standards</a><br /><b>Required action</b> - review IM status 'Pending client' and proceed according to <a href=https://skydrive3m.sharepoint.com/sites/PolCent/Policy%20Central/forms/DispForm.aspx?ID=10703&Source=https://skydrive3m.sharepoint.com/sites/PolCent&InitialTabID=Ribbon.Read%20>IT GSOP-040</a><br /> * * * * * * * * * * * * * * * * * * * *'''
fpath=sys.argv[1]
alphine_path=sys.argv[2]
mail_details_path=sys.argv[3]

folderpath=r"\\acprd01\E\3M_CAC\Alphine\textfiles"#sys.argv[1]

files= os.listdir(folderpath)
#print(files)
txtfile=files[0]


#cc & bcc
w=openpyxl.load_workbook(r"\\acprd01\E\3M_CAC\Alphine\Mail_Details\Alphine_mail_details.xlsx")#sys.argv[3])
s=w.worksheets[0]
c_c=","
b_c_c=","
for i in range(2,s.max_row+1):
    if s.cell(column=1,row=i).value not in [None,"NULL",""," "]:
        c_c+=str(s.cell(column=1,row=i).value.strip())+"@mmm.com,"
    if s.cell(column=2,row=i).value not in [None,"NULL",""," "]:
        b_c_c+=str(s.cell(column=2,row=i).value.strip())+"@mmm.com,"
c_c=c_c[:-1]
b_c_c=b_c_c[:-1]

def sentmail(textfile):
    recipient = textfile[:textfile.index('.txt')].split("_")
    recipient = list(filter(None, recipient))
    if len(recipient) == 1:
        to = recipient[0] + "@mmm.com"
        cc = ""
        bcc = ""
    elif len(recipient) == 2:
        to = recipient[0] + "@mmm.com"
        cc = recipient[1] + "@mmm.com"
        bcc = ""
    elif len(recipient) == 3:
        to = recipient[0] + "@mmm.com"
        cc = recipient[1] + "@mmm.com"
        bcc = recipient[2] + "@mmm.com"
    with open(folderpath+"\\"+textfile,"r",encoding='utf-16') as f:
        note_data=f.read()
    # Collecting the index
    incident_index = list()
    priority_in = list()
    Alert_ind = list()
    Required_action_ind = list()
    url_delimiter_open_l = list()
    url_delimiter_close_l = list()
    priority_delimiter_open_l = list()
    priority_delimiter_close_l = list()
    Alert_delimiter_open_l = list()
    Alert_delimiter_close_l = list()

    def index_collection(word, listname, x):
        for i in re.finditer(word, x):
            listname.append(i.start())

    index_collection('<b>Incident</b>', incident_index, note_data)
    index_collection('Priority', priority_in, note_data)
    index_collection('Alert', Alert_ind, note_data)
    index_collection('<b>Required action</b>', Required_action_ind, note_data)
    index_collection('url_delimiter_open', url_delimiter_open_l, note_data)
    index_collection('url_delimiter_close', url_delimiter_close_l, note_data)
    index_collection('priority_delimiter_open', priority_delimiter_open_l, note_data)
    index_collection('priority_delimiter_close', priority_delimiter_close_l, note_data)
    index_collection('Alert_delimiter_open', Alert_delimiter_open_l, note_data)
    index_collection('Alert_delimiter_close', Alert_delimiter_close_l, note_data)

    # print(url_delimiter_open_l,url_delimiter_close_l)

    # geting required data
    # Incident url and Incident ID
    incident_urls = list()
    incident_id = list()
    for i in range(0, len(url_delimiter_open_l)):
        batch = note_data[url_delimiter_open_l[i] + len('url_delimiter_open'):url_delimiter_close_l[i]]
        # cut = batch[batch.index("<a href"):batch.index("</a><br /> <b>") + 4].strip()
        incident_urls.append(batch)
        id = batch[batch.index("aspx?ID=") + len("aspx?ID="):batch.index("aspx?ID=") + len("aspx?ID=") + 10].strip()
        incident_id.append(id)

    # Fetching Priority
    priority_list = list()
    for i in range(0, len(url_delimiter_open_l)):
        pr_batch = note_data[
                   priority_delimiter_open_l[i] + len('priority_delimiter_open'):priority_delimiter_close_l[i]]
        # pr_cut = pr_batch[pr_batch.index("Priority</b> -") + len("Priority</b> -"):pr_batch.index("<br /><b>")].strip()
        priority_list.append(pr_batch)
    # print(Alert_ind, Required_action_ind)
    # Fetching Alert
    Alert_list = list()
    for i in range(0, len(url_delimiter_open_l)):
        al_batch = note_data[Alert_delimiter_open_l[i] + len('Alert_delimiter_open'):Alert_delimiter_close_l[i]]
        # al_cut = al_batch[al_batch.index("Alert</b> -") + len("Alert</b> -"):al_batch.index("<br /><b>")].strip()
        Alert_list.append(al_batch)
    # print(incident_urls,incident_id,priority_list,Alert_list)

    # Fetching Required Action
    req_action_list = list()
    for i in range(0, len(url_delimiter_open_l)):
        if i != len(url_delimiter_open_l) - 1:
            if 'please update incident log according' in note_data[Alert_delimiter_close_l[i]:url_delimiter_open_l[i + 1]]:
                req_action_list.append('please update incident log according to <a href=https://skydrive3m.sharepoint.com/sites/GOS/GOS%20Wiki/GOS%20Incident%20Support%20Process%20Guidelines/GOS%20Support%20Checklist.aspx?WikiPageMode=Edit&InitialTabId=Ribbon.EditingTools.CPEditTab&VisibilityContext=WSSWikiPage>3M Standards</a>')
            elif "review IM status 'Pending client' and proceed according to" in note_data[Alert_delimiter_close_l[i]:url_delimiter_open_l[i + 1]]:
                req_action_list.append("review IM status 'Pending client' and proceed according to <a href=https://sts.mmm.com/adfs/ls/?client-request-id=115639a0-a0e3-1000-b3c3-3030b4bc9fb2&username=&wa=wsignin1.0&wtrealm=urn%3afederation%3aMicrosoftOnline&wctx=estsredirect%3d2%26estsrequest%3drQQIARAAnVE_aBNRHL7XS69JbDUWBEeRTMpL3rv37t67gMN7dxcEtelglbqU-0uO9HLJJfXf6CROnRzq0OIYECE6SHEQ3DoFunVx6BIKgjp10wYXx-I3fPyGj4_v9303VFzDjSr6CwJnDFEcYxhEs-sf5Mvlyp714PLxnlL58BWhj593xztgIU3TWpClI2C0h8PeoFGvDzrPwjx5HJG0Nmh7edTLku5wpqlvxFmeDuphFHtbm8OaN-g9_QTABIApAKO5vmCmIIYjCbckI5JRYnOTCeRKalNiMZtSjAgVDhEmQqIJmSSCMoyxxVxCpGu7FhWs6UrsYOlw2qSCC8IcbBjSaRrMZMx1pS11adtnfubR3KWW2Bq29RllefI8-jVXmmXc6GWD4Y5aaDnoyUg9Vz1jtRp7gReQgMJI9wxIdYNBL2Ym1JERcI_rOAz9A1XLelE3CScFcFJYRGqjWCxXlKvKNeW0AN7On5X87udJ78r7b3d2D-9r7eOacjBfd-RD1-NrYjNbtzs3H91F7Xq_L1fypr6ei9xqreI1J7FWOiS9d4s38LYGtjVtXysV1YpyXbVX8VQDPzTwckHZL_3fTpML4GgRl0tB5udeN0zC5SrGfsgRJ5CzCEGKfQP6PuMQ-QEls4cj7r1YOpfsy5JyenH65vD1-Per77f_AA2#>IT GSOP-040</a>")
            elif "please work with Incident Manager and inform about next steps to close this incident" in note_data[Alert_delimiter_close_l[i]:url_delimiter_open_l[i + 1]]:
                req_action_list.append('please work with Incident Manager and inform about next steps to close this incident')
        else:
            if 'please update incident log according' in note_data[Alert_delimiter_close_l[i]:]:
                req_action_list.append('please update incident log according to <a href=https://skydrive3m.sharepoint.com/sites/GOS/GOS%20Wiki/GOS%20Incident%20Support%20Process%20Guidelines/GOS%20Support%20Checklist.aspx?WikiPageMode=Edit&InitialTabId=Ribbon.EditingTools.CPEditTab&VisibilityContext=WSSWikiPage>3M Standards</a>')
            elif "review IM status 'Pending client' and proceed according to" in note_data[Alert_delimiter_close_l[i]:]:
                req_action_list.append("review IM status 'Pending client' and proceed according to <a href=https://sts.mmm.com/adfs/ls/?client-request-id=115639a0-a0e3-1000-b3c3-3030b4bc9fb2&username=&wa=wsignin1.0&wtrealm=urn%3afederation%3aMicrosoftOnline&wctx=estsredirect%3d2%26estsrequest%3drQQIARAAnVE_aBNRHL7XS69JbDUWBEeRTMpL3rv37t67gMN7dxcEtelglbqU-0uO9HLJJfXf6CROnRzq0OIYECE6SHEQ3DoFunVx6BIKgjp10wYXx-I3fPyGj4_v9303VFzDjSr6CwJnDFEcYxhEs-sf5Mvlyp714PLxnlL58BWhj593xztgIU3TWpClI2C0h8PeoFGvDzrPwjx5HJG0Nmh7edTLku5wpqlvxFmeDuphFHtbm8OaN-g9_QTABIApAKO5vmCmIIYjCbckI5JRYnOTCeRKalNiMZtSjAgVDhEmQqIJmSSCMoyxxVxCpGu7FhWs6UrsYOlw2qSCC8IcbBjSaRrMZMx1pS11adtnfubR3KWW2Bq29RllefI8-jVXmmXc6GWD4Y5aaDnoyUg9Vz1jtRp7gReQgMJI9wxIdYNBL2Ym1JERcI_rOAz9A1XLelE3CScFcFJYRGqjWCxXlKvKNeW0AN7On5X87udJ78r7b3d2D-9r7eOacjBfd-RD1-NrYjNbtzs3H91F7Xq_L1fypr6ei9xqreI1J7FWOiS9d4s38LYGtjVtXysV1YpyXbVX8VQDPzTwckHZL_3fTpML4GgRl0tB5udeN0zC5SrGfsgRJ5CzCEGKfQP6PuMQ-QEls4cj7r1YOpfsy5JyenH65vD1-Per77f_AA2#>IT GSOP-040</a>")
            elif "please work with Incident Manager and inform about next steps to close this incident" in note_data[Alert_delimiter_close_l[i]:]:
                req_action_list.append("please work with Incident Manager and inform about next steps to close this incident")
    alphine = openpyxl.load_workbook(r"\\acprd01\E\3M_CAC\Alphine\alphine.xlsm")#sys.argv[2])
    alphinegroups = alphine.worksheets[0]
    gsr_update = alphine.worksheets[1]
    reportlog = alphine.worksheets[2]

    # Dictonaries
    Groups = dict()
    ownername = dict()
    opendays = dict()

    for i in range(2, gsr_update.max_row + 1):
        Groups[gsr_update.cell(column=39, row=i).value] = gsr_update.cell(column=4, row=i).value
        ownername[gsr_update.cell(column=39, row=i).value] = gsr_update.cell(column=6, row=i).value
        opendays[gsr_update.cell(column=39, row=i).value] = gsr_update.cell(column=48, row=i).value

    grouplist = list()
    ownerlist = list()
    opendayslists = list()
    for i in range(0, len(incident_id)):
        grouplist.append(Groups[incident_id[i]])
        ownerlist.append(ownername[incident_id[i]])
        opendayslists.append(opendays[incident_id[i]])
    # creating dataframe

    data = {'incident ID': incident_urls,'Support Groups': grouplist, 'Assignee': ownerlist, 'Open Days': opendayslists, 'Priority': priority_list, 'Alert': Alert_list,'Required Action':req_action_list}
    dframe = pd.DataFrame(data)
    #dframe.to_excel(folderpath+"\\"+"vbvb.xlsx")
   #Create HTML TABLE
    st='<tr style="background-color:#f2f2f2">'+"\n"
    sp='</tr>'+'\n'
    rs='<td class = "number_column">'
    rp='</td>'+'\n'
    hcode=str()
    for i in range(0,len(dframe)):
        hcode+=st
        for j in range(0, len(dframe.columns)):
            hcode+=rs+str(list(dframe.iloc[:, j])[i])+rp
        hcode+=sp

    #with open(folderpath+"\\"+"httmfile.txt","w") as g:
         #g.write(hcode)
    html_file = '''

    <!DOCTYPE html>
    <html>
    <head>
    <style>
    table {
      border-style:ridge;
      border-color:#000000;
      background-color:#000000;
      border= 1px solid;
      border-collapse: collapse;
      width: 100%;
    }

    table th {
      border : 1px solid #000000;
      padding: 6px;
      font-family: Helvetica, Arial, Helvetica;
      font-size: 12px;
      height: 30px;
    }

    table td{
      border : 1px solid #000000;
    }


    .header {
      color: white;
      background-color:#9400d3;
      border-bottom:1pt solid blue;
    }

    .text_column {
      text-align: centre;
      height:20px;
    }

    .number_column {
      text-align: centre;
      height:20px;
      width:100px";
    }

    .even_row {
      background-color: #f2f2f2;
    }

    </style>
    </head>

    <body>
      <h1></h1>
      <body style="font-family:Times New Roman">
      <br/><img src='cid:image1'<br/>
      <br>
      <br>
      <br /><font face='Times New Roman'><b><i>Hello,We have noticed that you are currently working on following incidents which requires your immediate action:</a></i></b></font><br/>
      <br /><font face='Times New Roman'><b><i>  </a></i></b></font><br/>
      <br>
      <br>


    <div style="overflow-x:auto;">

    <table>
      <thead>
        <tr class = "header">
          <th class = "text_column">''' +dframe.columns[0]+'''</th>
          <th class = "text_column">''' +dframe.columns[1]+'''</th>
          <th class = "text_column">''' +dframe.columns[2]+'''</th>
          <th class = "text_column">''' +dframe.columns[3]+'''</th>
          <th class = "text_column">''' +dframe.columns[4]+'''</th>
          <th class = "text_column">''' +dframe.columns[5]+'''</th>
          <th class = "text_column">''' +dframe.columns[6]+'''</th>
        </tr>
      </thead>
      <tbody>
    ''' + hcode+ '''
      </tbody>
      </table>

    </div>
    <br /><font face='Times New Roman'><b><i>Regards </a></i></b></font><br/>
    <br /><font face='Times New Roman'><b><i>3M Automation Center Team </a></i></b></font><br/>
    <br>
    <br>
    <br/><img src='cid:image3'<br/>

    </div>
    </body>
    </html>
    '''
    msgRoot = MIMEMultipart('related')
    msgRoot['Subject'] = "Incident requires your attention!"  #
    msgRoot['From'] = 'USSACPrd@mmm.com'#'USSACDev@mmm.com'
    msgRoot['Cc'] = c_c+","+cc
    msgRoot['To'] = to#"P.Anuraj@cognizant.com" #to
    msgRoot['Bcc'] = b_c_c+","+bcc
    msgRoot.preamble = '====================================================='
    msgAlternative = MIMEMultipart('alternative')
    msgRoot.attach(msgAlternative)
    msgText = MIMEText('Please find ')
    msgAlternative.attach(msgText)
    msgText = MIMEText(html_file, 'html')
    msgAlternative.attach(msgText)
    msgAlternative.attach(msgText)
    fp = open("//acprd01/E/3M_CAC/Alphine/Mail_image/head.png",
              'rb')  # "//acdev01/3M_CAC/IPM_FSM/Mail_elements/head.png"
    # fp2 = open(sys.argv[7], 'rb')#"//acdev01/3M_CAC/IPM_FSM/Mail_elements/new.png"
    fp3 = open("//acprd01/E/3M_CAC/Alphine/Mail_image/footer.png",
               'rb')  # "//acdev01/3M_CAC/IPM_FSM/Mail_elements/footer.png"
    msgImage = MIMEImage(fp.read())
    # msgImage1 = MIMEImage(fp2.read())
    msgImage2 = MIMEImage(fp3.read())
    fp.close()
    fp3.close()
    msgImage.add_header('Content-ID', '<image1>')
    msgImage2.add_header('Content-ID', '<image3>')
    msgRoot.attach(msgImage)
    msgRoot.attach(msgImage2)
    smtp = smtplib.SMTP()
    smtp.connect('mailserv.mmm.com')
    # smtp.sendmail(From,To, msgRoot.as_string())
    smtp.send_message(msgRoot)
    smtp.quit()
    print("Email is sent successfully", to, cc)

    #alphine.save(sys.argv[2])
    #alphine.close()
for i in files:
    sentmail(i)
#alphine.save(sys.argv[2])
#alphine.close()

#sentmail(txtfile)
print('success')

#-1 removed in range-->#for i in range(0, len(incident_index))
#encoding has added
#while creating data frame all rows must have same count ,other wise value error occurs

















