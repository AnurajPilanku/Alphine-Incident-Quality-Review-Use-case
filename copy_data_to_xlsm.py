import openpyxl
import sys

try:
    SHEET_POSITION = 1
    START_INDEX = 1

    inputFilePath = sys.argv[2]
    outputFilePath = sys.argv[1]
    sheetName = sys.argv[3]

    inputWorkbook = openpyxl.load_workbook(inputFilePath)
    inputSheet = inputWorkbook.active

    outputWorkbook = openpyxl.load_workbook(filename=outputFilePath, read_only=False, keep_vba=True)
    del outputWorkbook[sheetName]
    outputWorkbook.create_sheet(title=sheetName, index=SHEET_POSITION)
    outputSheet = outputWorkbook[sheetName]

    for row_num in range(START_INDEX, inputSheet.max_row+1):
        for col_num in range(START_INDEX, inputSheet.max_column+1):
            assignmentGroup = inputSheet.cell(row_num,4).value
            if(assignmentGroup!=None):
                cellValue = inputSheet.cell(row_num, col_num).value
                outputSheet.cell(row_num, col_num).value = cellValue
            else:
                break

    inputWorkbook.close()
    outputWorkbook.save(outputFilePath)
    outputWorkbook.close()

    print("success")

except Exception as e:
    print(e)
            
            
